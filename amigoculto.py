from openpyxl import load_workbook
from random import randint
from emailamigo import enviarmail

wb = load_workbook(r'C:\Users\Julio\Desktop\amigo_oculto.xlsx')
ws = wb['Planilha1']
dictdados = {}
listnames = []
numero_linha = ws.max_row
if (numero_linha - 1) % 2 != 0:
    print('Numero de participantes impar!!!')
    exit()
else:
    for linha in range(2, numero_linha + 1):
        nome = ws.cell(row=linha, column=2).value
        listnames.append(nome)
    for indice, user in enumerate(listnames):
        # indice = ws.cell(row=linha, column=1).value
        # nome = ws.cell(row=linha, column=2).value
        email = ws.cell(row=indice+2, column=3).value
        tupladados = (user, email)
        dictdados[indice+1] = tupladados
        dictdados.copy()

listamigos = []
flag = True
index = 1
while flag:
    amigoculto = randint(1, numero_linha - 1)
    if amigoculto in listamigos or amigoculto == index:
        pass
    else:
        dename = dictdados[index][0]
        fromadress = dictdados[index][1]
        destname = dictdados[amigoculto][0]
        # to = dictdados[amigoculto][1]
        listamigos.append(amigoculto)
        index += 1
        print(f'{dename} ({fromadress}) o seu amigo oculto é --> {destname}')
        enviarmail(fromadress, dename, destname)
    if index == numero_linha:
        flag = False
